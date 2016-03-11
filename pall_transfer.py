# This can be used to copy specific WT translations from one file to another
# You have to specify the name of the WT's tabs that correspond with each
# other in the separate files within the sheet_mapper dict
#
# I used this on Pall when PS QA and fixed some WT's by copying the WT's
# and fixing those, which did not carry the original WTs' translations
# over.  Since the titles of the WTs had changed and I only wanted to
# translate specific WT's I created the mapper
#
# Need to run this with 
# python multilang_format.py [name of template sheet] [name of error sheet]

import sys
import openpyxl

template_name = sys.argv[1]
errors_name = sys.argv[2]

# Mapper created as destination_file:source_file
sheet_mapper = {
 "How do I Change an As-steps(6)":"Recovered_Sheet1",
 "How do I Change an As-steps(2)":"Recovered_Sheet2",
 "How do I Change an As-steps(3)":"How do I change an as-steps(2)",
 "How do I View My Empl-steps":"How do I view my empl-steps",
 "How do I Add My Photo-steps":"How do I add my photo-steps",
 "How do I Initiate an -steps":"How do I initiate an -steps",
 "How do I View the Glo-steps":"How do I view the glo-steps",
 "How do I Change an As-steps(4)":"QAHow do I change an -steps",
 "How do I Update My De-steps":"How do I update my de-steps",
 "How do I View My Pers-steps":"How do I view my pers-steps",
 "How do I Change an As-steps(5)":"How do I change an as-steps(3)",
 "How do I Change an As-steps(7)":"How do I change an as-steps(4)"
}

template = openpyxl.load_workbook(template_name)
errors = openpyxl.load_workbook(errors_name)

template_sheets = template.get_sheet_names()
error_sheets = errors.get_sheet_names()

ntempsheets = len(template_sheets)

nerrorsheets = len(error_sheets)

# loop through each sheet, find one in the mapper
for n in range(ntempsheets):
    # Store current sheet template name, current template sheet
    temp_name_curr = template_sheets[n]
    temp_curr = template.get_sheet_by_name(temp_name_curr)

    if temp_name_curr in sheet_mapper:
        # loop through error file to find the template sheet's companion
        for x in range(nerrorsheets):

            # Store error sheet template name, error template sheet
            error_name_curr = error_sheets[x]
            error_curr = errors.get_sheet_by_name(error_name_curr)

            if sheet_mapper[temp_name_curr] == error_name_curr:
                # once the match is found, loop through each row to find matching text
                # in the error file
                for i in range(len(temp_curr.columns[0])-1):
                    temp_cell_title = temp_curr.cell(row=(i+2),column=2).value
                    temp_cell_descr = temp_curr.cell(row=(i+2),column=3).value

                    # once individual template cells stored, loop through error file
                    # to see if there's a match
                    for j in range(len(error_curr.columns[0])-1):
                        error_cell_title = error_curr.cell(row=(j+2),column=2).value
                        error_cell_descr = error_curr.cell(row=(j+2),column=3).value
                        # if a match in title or descr is found, store them
                        if error_cell_title == temp_cell_title:
                            temp_curr.cell(row=(i+2),column=4).value = error_curr.cell(row=(j+2),column=4).value
                        if error_cell_descr == temp_cell_descr:
                            temp_curr.cell(row=(i+2),column=5).value = error_curr.cell(row=(j+2),column=5).value

template.save('fixed_template.xlsx')