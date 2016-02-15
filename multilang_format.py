# Need to run this with 
# python multilang_format.py [name of template sheet] [name of error sheet]

import sys
import openpyxl

template_name = sys.argv[1]
errors_name = sys.argv[2]

template = openpyxl.load_workbook(template_name)
errors = openpyxl.load_workbook(errors_name)

template_sheets = template.get_sheet_names()
error_sheets = errors.get_sheet_names()

ntempsheets = len(template_sheets)

nerrorsheets = len(error_sheets)

if ntempsheets != nerrorsheets:
    print "Unequal number of sheets: template has " + str(ntempsheets) + ", error has " + str(nerrorsheets)
    extra_template = list(set(template_sheets) - set(error_sheets))
    extra_error = list(set(error_sheets) - set(template_sheets))
    if extra_template:
        print "The template has the following extra sheets: "
        for n in extra_template:
            print n
    if extra_error:
        print "The customer copy has the following extra sheets: "
        for n in extra_error:
            print n
    sys.exit()

for n in range(ntempsheets):
    # Set current sheet template name, current template sheet
    temp_name_curr = template_sheets[n]
    temp_curr = template.get_sheet_by_name(temp_name_curr)

    # Set error sheet template name, error template sheet
    error_name_curr = error_sheets[n]
    error_curr = errors.get_sheet_by_name(error_name_curr)

    if temp_name_curr != error_name_curr:
        error_curr.title = temp_name_curr

    # Fix the spreadsheet 1st row headers
    for x in range(len(temp_curr.rows[0])):
        error_cell = error_curr.cell(row=1,column=(x+1))
        temp_cell = temp_curr.cell(row=1,column=(x+1))
        if n == 0:
            print error_cell.value, temp_cell.value
        if error_cell.value != temp_cell.value:
            error_cell.value = temp_cell.value

errors.save('fixed_sheet.xlsx')